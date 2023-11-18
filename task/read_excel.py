

import json
import logging
import util.file_util as file_util
import pandas as pandas
import openpyxl
import numpy as numpy



file_name="D:\\jerry\\算法能力对比.xlsx"

#========================================使用pandas解析==============================================================

# content = pandas.read_excel(file_name,header=2,sheet_name="Sheet1")
# # print(content)
# print(list(content))
# # print(numpy.array(content)) 
# # print(content.get("序号"))
# list = []
# for row_list in numpy.array(content):
#     print(row_list)
#     dict = {
#             'algorithmName':row_list[1],
#             'algorithmId':row_list[2]
#             }
#     list.append(dict)

# # print(list)  
# print(type(json.dumps(list)))
# print(json.dumps(list,ensure_ascii=False))


wk = openpyxl.load_workbook(file_name)
sheet     = wk.get_sheet_by_name("Sheet1")
# print(sheet.max_row)
array = []
for r in range(0,sheet.max_row):
    row=[item.value for item in list(sheet.rows)[r]]
    # print(f'第{r}行值',row)
    dict = {
            'algorithmId':row[2],
            'resultNum':row[13],
            'accuracy':row[14]
            }
    array.append(dict)

d1 = {
    'provider':'002373',
    'algoList': array
}
print(json.dumps(d1))
