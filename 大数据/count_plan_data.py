'''
Author: error: error: git config user.name & please set dead value or install git && error: git config user.email & please set dead value or install git & please set dead value or install git
Date: 2024-06-17 10:01:30
LastEditors: error: error: git config user.name & please set dead value or install git && error: git config user.email & please set dead value or install git & please set dead value or install git
LastEditTime: 2024-06-18 08:58:38
FilePath: \python-learn\大数据\count_plan_data.py
Description: 这是默认设置,请设置`customMade`, 打开koroFileHeader查看配置 进行设置: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''
# 根据条件统计信号方案
import json
import logging
import util.file_util as file_util
import pandas as pandas
import openpyxl
import xlrd
import numpy as numpy

if __name__ == '__main__':
    file_name = "D:\data\CrossHxTimingPlan.xlsx"


    wk = openpyxl.load_workbook(file_name)
    sheet     = wk.get_sheet_by_name("CrossHxTimingPlan")
    print(sheet.max_row)
    array = []
    for r in range(1,sheet.max_row):
        row=[item.value for item in list(sheet.rows)[r]]
        row7_data = json.loads(row[7])
        for i in range(0,len(row7_data)):
            stageTime = row7_data[i]['stageTime']
            if int(stageTime) <= 11:
                # print(row7_data[i])
                dict = {
                    'crossNo':row[1],
                    'planNo':row[3],
                    'phaseNo': row7_data[i]['phaseNo'],
                    'stageTime': row7_data[i]['stageTime']
                }
                array.append(dict)
                file_util.append_to_file('D:\data\CrossHxTimingPlan3.json',str(dict)+'\n')  
        print(json.dumps(array))
 

    
    





        

    


