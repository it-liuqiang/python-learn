'''
Author: liuq liuq03@ehualu.com
Date: 2023-08-24 09:52:47
LastEditors: liuq liuq03@ehualu.com
LastEditTime: 2023-08-24 15:39:08
FilePath: \python-learn\read_excel.py
Description: 这是默认设置,请设置`customMade`, 打开koroFileHeader查看配置 进行设置: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''

import json
import logging
import util.file_util as file_util
import pandas as pandas
import openpyxl
import numpy as numpy



file_name="D:\\jerry\\事件码-字典&新增记录.xlsx"



content = pandas.read_excel(file_name,header=0,sheet_name="0-全目标")
# print(content)
print(list(content))
# print(numpy.array(content)) 
# print(content.get("序号"))
list = []
for row_list in numpy.array(content):
    print(row_list)
    dict = {'序号':row_list[0],
            '编码':row_list[1],
            '名称':row_list[2],
            '类型':row_list[3],
            '备注':row_list[4]
            }
    list.append(dict)

# print(list)  
print(type(json.dumps(list)))
print(json.dumps(list,ensure_ascii=False))


# wk = openpyxl.load_workbook(file_name)
# sheet     = wk.get_sheet_by_name("0-全目标")
# print(sheet.max_row)
# for r in range(0,sheet.max_row):
#     row=[item.value for item in list(sheet.rows)[r]]
#     print(f'第{r}行值',row)

